"""Read a workbook (Test4 format, read-only) into typed Python objects + a report.

The 3 master sheets (Machine / Operator & shift / Weekly off & holiday) are read
**header-driven** (columns located by heading, every row to the end) so Excel edits
flow in on re-upload with no code change. The SO list + process master keep their
column layout.

Design principles honoured here (CLAUDE.md):
  * The workbook is opened read-only; nothing is ever written back.
  * Loader-level data gaps are NON-BLOCKING: collect every problem into
    ``masters.report`` and keep going (PENDING_MASTER_DATA, NO_ROUTING, time
    coercions). The pipeline never stops here.

Sheet-name gotchas (trailing spaces, an apostrophe, a misspelling) are handled
by ``_find_sheet`` which matches on a normalized name.
"""
from __future__ import annotations

import math
import re
from datetime import date, datetime
from functools import lru_cache

import openpyxl

from .models import (
    SOLine,
    Process,
    Routing,
    Machine,
    Operator,
    WorkCalendar,
    Masters,
)

# Number of process blocks in the routing sheet, and the 5 columns per block.
MAX_PROCESSES = 12
ROUTING_FIRST_PROCESS_COL = 12  # 0-based col of "Process 1"


# --------------------------------------------------------------------------- #
# Normalization helpers
# --------------------------------------------------------------------------- #
@lru_cache(maxsize=None)
def _normalize_resource_id_cached(raw: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", raw.upper())


def normalize_resource_id(raw) -> str:
    """Canonical machine/resource id: uppercase, alnum only.

    Collapses the master's spaced labels ('CNC 4', 'VMC 1') and the routing's
    compact labels ('CNC4', 'VMC1') onto the same key, so they match without a
    hand-maintained alias table.

    Memoized: this is a pure string→string function called millions of times per
    optimizer run over a tiny set of distinct machine labels, so the regex runs
    once per distinct input, not once per call. Correctness is unchanged.
    """
    if raw is None:
        return ""
    return _normalize_resource_id_cached(raw if isinstance(raw, str) else str(raw))


def normalize_process_name(raw) -> str:
    """Canonical process-name key: collapse internal/leading/trailing whitespace and
    uppercase, but KEEP word breaks ('cnc  first side ' -> 'CNC FIRST SIDE'). Used to
    match a punched-in process to its routing step and finished-goods gate without
    being tripped by spacing/case (unlike ``normalize_resource_id``, which strips
    spaces entirely for machine ids)."""
    return " ".join(str(raw or "").split()).upper()


@lru_cache(maxsize=None)
def _parse_resource_candidates_cached(raw: str) -> tuple:
    out = []
    for token in re.split(r"[/&,]| or ", raw):
        cid = normalize_resource_id(token)
        if cid and cid not in out:
            out.append(cid)
    return tuple(out)


def parse_resource_candidates(raw) -> list:
    """Canonical machine ids a routing cell allows, in order (first = preferred).

    A "Suggested M/c" cell may list ALTERNATIVES separated by '/', ',', '&' or ' or '
    (e.g. 'CNC3/CNC6' = run on either CNC3 or CNC6). Returns an ordered, deduped list
    of normalized ids; empty/None -> []. The same split is used for operators.

    Memoized on the (fixed) cell text, so the split/normalize runs once per distinct
    routing cell instead of millions of times across an optimizer run. Returns a
    fresh list each call (callers may treat it as owned), built cheaply from the
    cached tuple — the regex work is what's saved."""
    if raw is None:
        return []
    return list(_parse_resource_candidates_cached(raw if isinstance(raw, str) else str(raw)))


def parse_date(value):
    """Coerce a cell to a ``date``. Real date/datetime cells are read by value
    (any Excel display format is fine). Text cells are parsed **strictly day-first**
    (DD/MM/YYYY, DD-MM-YYYY) or ISO — matching the app-wide DD-MM-YYYY policy, so an
    ambiguous string like '03/04/2025' is always 3 April, never 4 March. A
    month-first US string is rejected (returns None) rather than silently coerced."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):   # day-first + ISO; no %m/%d/%Y
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _norm_sheet(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(name).lower())


def _find_sheet(wb, wanted: str):
    """Locate a sheet by normalized name (tolerates trailing spaces, the
    apostrophe in "Item's process Master", and the 'anayasis' misspelling)."""
    target = _norm_sheet(wanted)
    for ws in wb.worksheets:
        if _norm_sheet(ws.title) == target:
            return ws
    # Fall back to a contains-match for the misspelled analysis sheet etc.
    for ws in wb.worksheets:
        if target in _norm_sheet(ws.title) or _norm_sheet(ws.title) in target:
            return ws
    return None


def _num(value, masters: Masters = None, ref: str = ""):
    """Coerce a time/number cell to a FINITE float; log a coercion if it was a string.
    Non-finite values (a cell typed 'nan'/'inf', or a NaN/inf float) are rejected as
    None — a non-finite cycle time would make occupancy NaN/inf and hang the scheduler."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value) if math.isfinite(value) else None
    try:
        coerced = float(str(value).strip())
    except (ValueError, TypeError):
        return None
    if not math.isfinite(coerced):
        return None
    if masters is not None:
        masters.add_report("TIME_COERCION", ref, f"coerced {value!r} -> {coerced}")
    return coerced


# --------------------------------------------------------------------------- #
# Header-driven table location (Test4 format)
# --------------------------------------------------------------------------- #
# The three master sheets are clean tables: a header row, then one row per record,
# open-ended. We locate columns by HEADING (not fixed positions) and read every
# row to the end — so adding/editing/removing rows in Excel "just works" on the
# next upload (no code change). See the master-data design note.
WEEKDAY_NAMES = {
    "monday": 0, "tuesday": 1, "wednesday": 2, "thursday": 3,
    "friday": 4, "saturday": 5, "sunday": 6,
}


def _norm_header(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).lower()) if value is not None else ""


def _locate_table(rows, needed: dict, exact_priority: bool = False):
    """Find the header row and map each canonical key to its column index.

    ``needed`` maps key -> a token that must appear in the normalized header cell
    (e.g. ``{"no": "machineno"}``). Position-independent. Returns
    ``(header_index, {key: col_index})`` or ``(None, {})`` if not all found.

    ``exact_priority`` (opt-in, default off ⇒ existing callers byte-identical):
    when True, each header row is matched in TWO passes — first bind any key whose
    normalized header cell EQUALS its token, then fill the still-unbound keys by the
    original leftmost-substring rule. This stops a token that is a substring of a
    neighbouring header from shadowing an exact column (e.g. "soqty" ⊂ "pendsoqty":
    the real "SO Qty" column wins regardless of whether "Pend SO Qty" sits to its
    left)."""
    for idx, row in enumerate(rows):
        colmap = {}
        if exact_priority:
            # Pass 1: exact equality wins (leftmost exact match per key).
            for ci, cell in enumerate(row):
                h = _norm_header(cell)
                if not h:
                    continue
                for key, token in needed.items():
                    if key not in colmap and h == token:
                        colmap[key] = ci
        # Substring pass (the original rule) fills whatever is still unbound.
        for ci, cell in enumerate(row):
            h = _norm_header(cell)
            if not h:
                continue
            for key, token in needed.items():
                if key not in colmap and token in h:
                    colmap[key] = ci
        if len(colmap) == len(needed):
            return idx, colmap
    return None, {}


def _cell(row, idx):
    return row[idx] if idx is not None and idx < len(row) else None


def _weekday_from_text(value):
    if value is None:
        return None
    low = str(value).lower()
    for name, wd in WEEKDAY_NAMES.items():
        if name in low:
            return wd
    return None


# --------------------------------------------------------------------------- #
# Individual sheet loaders
# --------------------------------------------------------------------------- #
def _load_machines(wb, masters: Masters):
    ws = _find_sheet(wb, "Machine master")
    if ws is None:
        masters.add_report("MISSING_SHEET", "Machine master", "sheet not found")
        return
    rows = list(ws.iter_rows(values_only=True))
    hdr, col = _locate_table(rows, {"type": "machinetype", "no": "machineno", "rate": "hrrate"})
    if hdr is None:
        masters.add_report("MISSING_SHEET", "Machine master",
                           "header row (Machine Type / Machine No / Hr Rate) not found")
        return
    # Available Hrs/Day is optional — locate it within the header row if present.
    avail_col = next((ci for ci, c in enumerate(rows[hdr])
                      if "availablehrs" in _norm_header(c)), None)
    for row in rows[hdr + 1:]:
        machine_no = _cell(row, col["no"])
        if not machine_no or str(machine_no).strip() == "":
            continue  # skip blank rows (open-ended table)
        canonical = normalize_resource_id(machine_no)
        if not canonical:
            continue
        machine_type = _cell(row, col["type"])
        masters.machines[canonical] = Machine(
            machine_no=canonical,
            display_name=str(machine_no).strip(),
            machine_type=str(machine_type).strip() if machine_type else "",
            hr_rate=_num(_cell(row, col["rate"])),
            provisional=False,
            available_hrs_per_day=(_num(_cell(row, avail_col)) if avail_col is not None else None),
        )


def _load_operators(wb, masters: Masters):
    ws = _find_sheet(wb, "Operator & shift Master")
    if ws is None:
        return
    rows = list(ws.iter_rows(values_only=True))
    hdr, col = _locate_table(rows, {"name": "operatorname", "pref": "preferredmachine"})
    if hdr is None:
        return  # no operators table (non-blocking)
    # The operator's own Shift column sits immediately after Preferred Machines.
    # (A separate shift-definitions table further right also has a "Shift" header —
    # bind only the adjacent one so the two don't collide.)
    shift_col = col["pref"] + 1
    if "shift" not in _norm_header(_cell(rows[hdr], shift_col)):
        shift_col = None
    for row in rows[hdr + 1:]:
        name = _cell(row, col["name"])
        if not name or str(name).strip() == "":
            continue  # blank row (the side-by-side shift table has no operator name)
        pref = _cell(row, col["pref"])
        parsed = [normalize_resource_id(t) for t in re.split(r"[/&,]| or ", str(pref or ""))]
        parsed = [p for p in parsed if p]
        shift = _cell(row, shift_col) if shift_col is not None else None
        masters.operators.append(
            Operator(name=str(name).strip(),
                     preferred_machines_raw=str(pref or "").strip(), machines=parsed,
                     shift=str(shift).strip() if shift else "")
        )


def _load_calendar(wb, masters: Masters):
    ws = _find_sheet(wb, "Weekly off & holiday master")
    cal = WorkCalendar()
    if ws is None:
        masters.calendar = cal
        return
    rows = list(ws.iter_rows(values_only=True))
    hdr, col = _locate_table(rows, {"cat": "category", "name": "name", "date": "date"})
    if hdr is None:
        masters.add_report("MISSING_SHEET", "Weekly off & holiday master",
                           "header row (Category / Name / Day-Date) not found")
        masters.calendar = cal
        return
    for row in rows[hdr + 1:]:
        cat = _cell(row, col["cat"])
        if not cat or str(cat).strip() == "":
            continue
        category = str(cat).strip().lower()
        name = _cell(row, col["name"])
        when = _cell(row, col["date"])
        if category.startswith("weekly"):
            # Weekly off day comes from the date cell ("Every Thursday") or the name.
            wd = _weekday_from_text(when)
            if wd is None:
                wd = _weekday_from_text(name)
            if wd is not None:
                cal.weekly_off_weekday = wd
        elif category.startswith("holiday"):
            d = parse_date(when)
            if d is not None:
                cal.holidays.append(d)
        elif category.startswith("leave"):
            d = parse_date(when)
            if d is not None:
                cal.leaves.append((str(name).strip() if name else "", d))
    masters.calendar = cal


def _load_routings(wb, masters: Masters):
    ws = _find_sheet(wb, "Item's process Master")
    if ws is None:
        masters.add_report("MISSING_SHEET", "Item's process Master", "sheet not found")
        return
    for row in ws.iter_rows(min_row=3, values_only=True):
        item_code = _cell(row, 3)
        if item_code is None or str(item_code).strip() == "":
            continue  # blank / separator row
        code = str(item_code).strip()
        processes = []
        for p in range(MAX_PROCESSES):
            base = ROUTING_FIRST_PROCESS_COL + p * 5
            name = _cell(row, base)
            if not name or str(name).strip() == "":
                continue
            sm, am = _cell(row, base + 3), _cell(row, base + 4)
            processes.append(
                Process(
                    seq=p + 1,
                    name=str(name).strip(),
                    cycle_time=_num(_cell(row, base + 1), masters, f"{code} P{p+1} cycle"),
                    total_time=_num(_cell(row, base + 2), masters, f"{code} P{p+1} total"),
                    suggested_machine=(str(sm).strip() if sm else None),
                    allotted_machine=(str(am).strip() if am else None),
                )
            )
        # Per-process progress (WIP 'continue from reality' re-planning) is keyed by the
        # NORMALISED process name, so two steps in one routing that normalise to the same
        # name would merge — their completed/remaining qty would be wrong and could exceed
        # the ordered qty. Owner-confirmed this never happens in the current data; report
        # it loudly (non-blocking) if a future workbook ever introduces it.
        _by_name: dict = {}
        for pr in processes:
            _by_name.setdefault(normalize_process_name(pr.name), []).append(pr.seq)
        for _nm, _seqs in _by_name.items():
            if len(_seqs) > 1:
                masters.add_report(
                    "DUPLICATE_PROCESS", str(code),
                    f"process name {_nm!r} appears at steps {_seqs}: per-step progress "
                    f"would merge — give each step a distinct name")

        masters.routings[code] = Routing(
            item_code=code,
            description=str(_cell(row, 2)).strip() if _cell(row, 2) else "",
            customer=str(_cell(row, 1)).strip() if _cell(row, 1) else "",
            rm_type=str(_cell(row, 6)).strip() if _cell(row, 6) else "",
            moq=_num(_cell(row, 10)),
            processes=processes,
        )


def _load_so_lines(wb, masters: Masters):
    """Read the SO list header-driven (position-independent), like the 3 master
    sheets — the owner's ERP export shifts these columns on every export, and
    NOT uniformly, so a fixed-index reader silently reads the wrong columns
    (e.g. delivery date lands on a text cell) and drops every row."""
    ws = _find_sheet(wb, "Sales Order (SO) list")
    if ws is None:
        masters.add_report("MISSING_SHEET", "Sales Order (SO) list", "sheet not found")
        return []
    rows = list(ws.iter_rows(values_only=True))
    hdr, col = _locate_table(rows, {
        "so_no": "sono",
        "item_code": "salesitemcode",
        "qty": "soqty",
        "delivery": "sodeliverydate",
    }, exact_priority=True)   # "soqty" ⊂ "pendsoqty" — exact "SO Qty" must win
    if hdr is None:
        masters.add_report(
            "MISSING_SO_COLUMNS", "Sales Order (SO) list",
            "could not find the SO number / item code / quantity / delivery date "
            "columns by header — check the sheet headers",
        )
        return []
    # Optional columns: bind if present in the same header row (first cell wins,
    # left to right), else leave unbound so _cell reads None. Not folded into
    # `_locate_table`'s `needed` dict — that would make them required, and the
    # real workbook has no "Remarks" column at all.
    header_row = rows[hdr]
    for key, token in {
        "item_name": "salesitemname",
        "customer": "customername",
        "pending_qty": "pendsoqty",
        "remarks": "remarks",
    }.items():
        for ci, cell in enumerate(header_row):
            h = _norm_header(cell)
            if h and token in h:
                col[key] = ci
                break

    so_lines = []
    for row in rows[hdr + 1:]:
        item_code = _cell(row, col.get("item_code"))
        if item_code is None or str(item_code).strip() == "":
            continue
        so_no = _cell(row, col.get("so_no"))
        delivery_raw = _cell(row, col.get("delivery"))
        delivery = parse_date(delivery_raw)
        if delivery is None:
            masters.add_report(
                "BAD_DELIVERY_DATE", str(so_no), f"unparseable delivery date {delivery_raw!r}"
            )
            continue
        item_name = _cell(row, col.get("item_name"))
        customer = _cell(row, col.get("customer"))
        remarks = _cell(row, col.get("remarks"))
        qty_raw = _cell(row, col.get("qty"))
        qty_val = _num(qty_raw)
        if qty_val is None and qty_raw is not None and str(qty_raw).strip() != "":
            masters.add_report(
                "BAD_QTY", str(so_no), f"unparseable SO qty {qty_raw!r}"
            )
        so_lines.append(
            SOLine(
                so_no=str(so_no).strip() if so_no else "",
                item_code=str(item_code).strip(),
                item_name=str(item_name).strip() if item_name else "",
                qty=qty_val or 0.0,
                delivery_date=delivery,
                pending_qty=_num(_cell(row, col.get("pending_qty"))),
                customer=str(customer).strip() if customer else "",
                remarks=str(remarks).strip() if remarks else "",
            )
        )
    return so_lines


# --------------------------------------------------------------------------- #
# Validation (non-blocking)
# --------------------------------------------------------------------------- #
def _register_provisional(masters: Masters, raw_label: str):
    """Ensure a resource referenced by a routing exists. If it isn't in the
    Machine master, register it as a PROVISIONAL machine and report it once."""
    canonical = normalize_resource_id(raw_label)
    if not canonical or canonical == "OS":
        return   # 'OS' marks outsourcing, not a machine — never a provisional resource
    if canonical in masters.machines:
        return
    masters.machines[canonical] = Machine(
        machine_no=canonical,
        display_name=str(raw_label).strip(),
        machine_type="(provisional: fill in Machine master)",
        hr_rate=None,
        provisional=True,
    )
    masters.add_report(
        "PENDING_MASTER_DATA",
        canonical,
        f"resource '{raw_label}' used by a routing but not in Machine master; "
        f"registered as provisional: add it to the Excel master to complete it",
    )


def _validate(masters: Masters, so_lines):
    # PENDING_MASTER_DATA: every resource a routing names must resolve.
    for routing in masters.routings.values():
        for proc in routing.processes:
            for raw in (proc.suggested_machine, proc.allotted_machine):
                # A cell may list alternatives ("CNC3/CNC6") — register EACH so a
                # missing one (e.g. VMC3) becomes its own provisional machine, not a
                # merged bogus id.
                for cid in parse_resource_candidates(raw):
                    _register_provisional(masters, cid)

    # NO_ROUTING: SO item codes with no recipe — report and (caller) skip them.
    seen = set()
    for so in so_lines:
        if so.item_code not in masters.routings and so.item_code not in seen:
            seen.add(so.item_code)
            masters.add_report(
                "NO_ROUTING",
                so.item_code,
                f"SO item '{so.item_code}' has no routing in Item's process Master; "
                f"order skipped (cannot schedule without a recipe)",
            )


# --------------------------------------------------------------------------- #
# Public entry point
# --------------------------------------------------------------------------- #
def load_all(xlsx_path):
    """Load masters + SO lines from a workbook in the Test4 format.

    ``xlsx_path`` is a path or a file-like object (e.g. an uploaded BytesIO) — there
    is no bundled default; callers always supply the source (an uploaded workbook,
    or the generated test sample).

    Returns ``(so_lines, masters)``. ``masters.report`` holds all non-blocking
    issues. SO lines whose item has no routing are dropped from the returned
    list (recorded as NO_ROUTING) so downstream rules only see schedulable demand.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    masters = Masters()
    try:
        _load_machines(wb, masters)
        _load_operators(wb, masters)
        _load_calendar(wb, masters)
        _load_routings(wb, masters)
        so_lines = _load_so_lines(wb, masters)
    finally:
        wb.close()

    _validate(masters, so_lines)

    schedulable = [so for so in so_lines if so.item_code in masters.routings]
    return schedulable, masters
