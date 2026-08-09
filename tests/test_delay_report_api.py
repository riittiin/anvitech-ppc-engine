"""Delay report — the xlsx serializer and the admin-only /delay-report.xlsx endpoint."""
import importlib
import io
from datetime import date

import pytest

pytest.importorskip("openpyxl")
pytest.importorskip("fastapi")
import openpyxl
from fastapi.testclient import TestClient

from engine import book_store
from engine.models import Order
from tests.sample_workbook import build_sample_bytes, ITEM_A


def _api():
    import api.main as m
    importlib.reload(m)
    return m


# ---- Task 5: serializer --------------------------------------------------- #

def test_delay_xlsx_has_two_sheets_and_headers():
    m = _api()
    report = {"summary": [{"SO No": "SO1", "Item Code": "X", "Item Name": "X",
              "Ordered Qty": 1, "SO Delivery Date": date(2025, 3, 20),
              "Expected Completion": date(2025, 3, 25), "Days Late": 5,
              "Working (days)": 1.0, "Waiting: machine (days)": 4.0,
              "Waiting: off-hours (days)": 0.0, "Waiting: crew (days)": 0.0,
              "Why": "5 days late — 4.0d machines busy"}],
              "detail": []}
    wb = openpyxl.load_workbook(io.BytesIO(m._delay_report_xlsx(report)))
    assert wb.sheetnames == ["Summary", "Detail"]
    assert wb["Summary"].cell(1, 1).value == "SO No"
    assert wb["Summary"].cell(2, 1).value == "SO1"


# ---- Task 6: endpoint ----------------------------------------------------- #

def test_delay_report_returns_xlsx_to_either_role():
    """Role-open since 2026-08-09 (director asked the two portals to match).

    Deliberate change, not a relaxed assertion: the report is a read-only view of
    the same plan both roles already see on Schedule and Gantt. The user-role leg
    below used to assert 403. Role PARITY is pinned in tests/test_role_parity.py,
    which also pins what stays admin-only (e.g. /efficiency).
    """
    m = _api()
    book_store.save_masters_bytes(build_sample_bytes())
    book_store.add_orders([Order("SO1", ITEM_A, ITEM_A, 40, date(2025, 3, 20))])
    m._current_masters()

    admin = TestClient(m.app)
    admin.post("/login", data={"username": "anvitech", "password": "1930rail"})
    r = admin.get("/delay-report.xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Summary", "Detail"]

    user = TestClient(m.app)
    user.post("/login", data={"username": "anvitech_user", "password": "anvitech12345678"})
    ru = user.get("/delay-report.xlsx")
    assert ru.status_code == 200, ru.text
    assert "spreadsheetml" in ru.headers["content-type"]
