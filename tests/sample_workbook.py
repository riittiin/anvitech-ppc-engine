"""Generate a tiny workbook in the Test3 format, in memory, for tests + golden.

Replaces the old bundled Test2.xlsx. Built from code each run — no committed
binary, no real data. It deliberately exercises the loader's behaviours:
  * two items with routings (so consolidation/priority/allocation run),
  * an ALTERNATIVE-machine cell ("CNC1/CNC2", both real),
  * a PROVISIONAL machine ("CNC9", referenced by a routing but not in the master),
  * a Thursday weekly-off + a holiday + an operator leave,
  * two same-item SO lines within the 10-day window (so Rule 1 consolidates them).

Sheet layouts:
  * the 3 master sheets use the new header-driven Test3 format (header row 1),
  * "Item's process Master" and "Sales Order (SO) list" use the same column
    positions as the real workbook (the loader was already built for those).
"""
import datetime
import io

import openpyxl

# --- known sample values (assertions in the tests reference these) --------- #
MACHINES = [
    ("CNC lathe", "CNC 1", 250, 19.5),
    ("CNC lathe", "CNC 2", 250, 19.5),
    ("Vertical Machining center", "VMC 1", 500, 19.5),
    ("Band saw cutting machine", "BS1", 100, 9.5),
    ("Manual Inspection", "MI1", 150, 9.5),
    ("Manual Washing", "MW1", 80, 9.5),
]  # 6 real machines

HOLIDAY = datetime.date(2025, 1, 1)
LEAVE_DATE = datetime.date(2025, 3, 12)
A_THURSDAY = datetime.date(2025, 3, 13)   # a normal working-calendar Thursday

# SO numbers / items the tests use.
SO1, SO2, SO3 = "SO-001", "SO-002", "SO-003"
ITEM_A, ITEM_B = "SAMP-A-01", "SAMP-B-01"


def _set(ws, row, mapping):
    """Set cells from a {0-based column index: value} mapping on ``row``."""
    for ci, val in mapping.items():
        ws.cell(row=row, column=ci + 1, value=val)


def build_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    # --- Machine master (new format: header row 1) --- #
    ws = wb.create_sheet("Machine master")
    ws.append(["Machine Type", "Machine No", "Hr Rate (Rs)", "Available Hrs/Day"])
    for r in MACHINES:
        ws.append(list(r))

    # --- Operator & shift Master (operators A:B:C, shift defs E:H side by side) --- #
    ws = wb.create_sheet("Operator & shift Master")
    ws.append(["Operator Name", "Preferred Machines", "Shift", None,
               "Shift", "Start", "End", "Notes"])
    # CNC1: covered First (One) + Second (Two) -> both shifts. CNC2: First only.
    # BS1 (a 9.5 machine): covered First (Three) -> runs the manual 9-6 window.
    ops = [
        ("Operator One", "CNC 1, CNC 2", "First shift"),
        ("Operator Two", "CNC 1", "Second shift"),
        ("Operator Three", "BS1", "First shift"),
    ]
    shifts = [("First shift", "08:00", "19:00", ""),
              ("Second shift", "19:00", "05:00", "next day")]
    for i in range(max(len(ops), len(shifts))):
        row = [None] * 8
        if i < len(ops):
            row[0], row[1], row[2] = ops[i]
        if i < len(shifts):
            row[4], row[5], row[6], row[7] = shifts[i]
        ws.append(row)

    # --- Weekly off & holiday master (Category | Name | Day / Date) --- #
    ws = wb.create_sheet("Weekly off & holiday master")
    ws.append(["Category", "Name", "Day / Date"])
    ws.append(["Weekly Off", "", "Every Thursday"])
    ws.append(["Holiday", "New Year", HOLIDAY])
    ws.append(["Leave", "Worker One", LEAVE_DATE])

    # --- Item's process Master (header row 2; data row 3+) --- #
    ws = wb.create_sheet("Item's process Master")
    _set(ws, 1, {5: "Order Type", 6: "Raw material data", 11: "Process sequence"})
    _set(ws, 2, {
        0: "Master Sr NO", 1: "Customer", 2: "Item Description", 3: "Item code",
        6: "RM type", 10: "MOQ/Batch qty",
        12: "Process 1", 13: "Process 1 cycle time", 14: "Process 1 Total time",
        15: "Process 1 Suggested M/c", 16: "Process 1 Allotted M/c",
        17: "Process 2", 18: "Process 2 cycle time", 19: "Process 2 Total time",
        20: "Process 2 Suggested M/c", 21: "Process 2 Allotted M/c",
        22: "Process 3", 23: "Process 3 cycle time", 24: "Process 3 Total time",
        25: "Process 3 Suggested M/c", 26: "Process 3 Allotted M/c",
    })
    # Item A: BANDSAW(BS1) -> CNC OS(CNC1/CNC2 alternative) -> INSP(MI1)
    _set(ws, 3, {
        0: 1, 1: "ALFA", 2: "SAMPLE RING A", 3: ITEM_A, 6: "Dia 20 SS", 10: 50,
        12: "BANDSAW", 13: 3, 14: 3, 15: "BS1",
        17: "CNC OS", 18: 5, 19: 5, 20: "CNC1/CNC2",
        22: "INSP", 23: 2, 24: 2, 25: "MI1",
    })
    # Item B: CNC(CNC9 = provisional) -> WASHING(MW1)
    _set(ws, 4, {
        0: 2, 1: "ALFA", 2: "SAMPLE PIN B", 3: ITEM_B, 6: "Dia 14 SS", 10: 100,
        12: "CNC", 13: 4, 14: 4, 15: "CNC9",
        17: "WASHING", 18: 2, 19: 2, 20: "MW1",
    })

    # --- Sales Order (SO) list (header row 1; data row 2+) --- #
    ws = wb.create_sheet("Sales Order (SO) list")
    _set(ws, 1, {5: "SONo", 8: "Customer Name", 19: "Sales Item Code",
                 20: "Sales Item Name", 21: "SO Qty", 23: "SO Delivery Date",
                 24: "Remarks", 27: "Pend SO Qty"})
    _set(ws, 2, {5: SO1, 8: "ALFA LAVAL", 19: ITEM_A, 20: "SAMPLE RING A",
                 21: 5, 23: datetime.date(2025, 3, 10), 24: "club with SO-002", 27: 5})
    _set(ws, 3, {5: SO2, 8: "ALFA LAVAL", 19: ITEM_A, 20: "SAMPLE RING A",
                 21: 10, 23: datetime.date(2025, 3, 14), 24: "", 27: 10})
    _set(ws, 4, {5: SO3, 8: "ALFA LAVAL", 19: ITEM_B, 20: "SAMPLE PIN B",
                 21: 4, 23: datetime.date(2025, 3, 20), 24: "", 27: 4})
    return wb


def build_sample_bytes() -> bytes:
    buf = io.BytesIO()
    build_workbook().save(buf)
    return buf.getvalue()
