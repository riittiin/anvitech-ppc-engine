"""Task 2 wiring: the app-owned operator table overlaid onto masters at ONE
point (display as-of today, planning as-of the plan's effective start), seeded
once from the workbook, carried through the cloud payload, folded into the
inputs fingerprint, and honoured by the scheduled-skip check.

The load-bearing invariants (spec 2026-07-18):
  1. A seeded-from-sample table plans BYTE-IDENTICALLY to the Excel-loaded plan.
  2. A re-upload NEVER mutates a seeded table.
  3. Planning rotates AS OF the plan's effective start (Friday shift-1 rule).
  4. Cloud (payload-carried table) == local (same table) byte-for-byte.
"""
import io
import json
from dataclasses import replace
from datetime import date

import pytest

from engine import book_store, operator_master
from engine import optimize_service as svc
from engine import optimizer, orderbook
from engine.config import Config, OVERLAP_PERCENT
from engine.loaders import load_all
from engine.models import Order, PlanRun
from engine.pipeline import run_forward
from tests.sample_workbook import build_sample_bytes

pytest.importorskip("fastapi")


def _api():
    import importlib
    import api.main as m
    importlib.reload(m)
    return m


def _orders(so_lines):
    orders = {}
    for so in so_lines:
        o = Order(so.so_no, so.item_code, so.item_name, so.qty, so.delivery_date)
        orders[o.key] = o
    return orders


def _cfg(start=date(2025, 3, 1), overlap=80):
    c = Config(overlap_mode=OVERLAP_PERCENT, overlap_percent=overlap,
               plan_start_date=start)
    c.validate()
    return c


def _workbook_with_changed_operator():
    """Sample workbook whose Operator sheet names a DIFFERENT first operator —
    used to prove a re-upload never rewrites a seeded table."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(build_sample_bytes()))
    ws = wb["Operator & shift Master"]
    ws.cell(row=2, column=1, value="CHANGED PERSON")   # was "Operator One"
    ws.cell(row=2, column=3, value="Second shift")      # was "First shift"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# --------------------------------------------------------------------------- #
# 1. Seeded == Excel-loaded (byte-identical schedule).
# --------------------------------------------------------------------------- #
def test_seeded_sample_plan_is_byte_identical_to_excel_loaded():
    raw = build_sample_bytes()
    so_lines, excel_masters = load_all(io.BytesIO(raw))
    cfg = _cfg()

    pr_excel = PlanRun(so_lines=list(so_lines))
    run_forward(pr_excel, cfg, excel_masters)

    m = _api()
    book_store.save_masters_bytes(raw)
    seeded_masters = m._current_masters()          # seeds + overlays operators
    pr_seed = PlanRun(so_lines=list(so_lines))
    run_forward(pr_seed, cfg, seeded_masters)

    assert [e.as_row() for e in pr_seed.schedule] == \
           [e.as_row() for e in pr_excel.schedule]
    # And the store was seeded exactly once.
    assert book_store.load_operator_table() is not None


# --------------------------------------------------------------------------- #
# 2. Upload never mutates a seeded table.
# --------------------------------------------------------------------------- #
def test_upload_never_mutates_a_seeded_table():
    m = _api()
    book_store.save_masters_bytes(build_sample_bytes())
    m._current_masters()                            # seed from A
    table_a = book_store.load_operator_table()
    names_a = {r["name"] for r in table_a["operators"]}
    assert "Operator One" in names_a

    # Re-upload a workbook with a different operator sheet.
    book_store.save_masters_bytes(_workbook_with_changed_operator())
    m._MASTERS_CACHE["masters"] = None              # force a re-read (upload does this)
    masters_after = m._current_masters()

    assert book_store.load_operator_table() == table_a          # store untouched
    assert {o.name for o in masters_after.operators} == names_a  # display still A
    assert "CHANGED PERSON" not in {o.name for o in masters_after.operators}


# --------------------------------------------------------------------------- #
# 3. Planning rotates AS OF the effective plan start (Friday shift-1 rule).
# --------------------------------------------------------------------------- #
def _shift_of(masters, name):
    return next(o.shift for o in masters.operators if o.name == name)


def test_prepare_contest_uses_the_operator_table_and_never_mutates_it():
    """prepare_contest overlays the app-owned operator table onto masters, in place
    of the workbook's operators — and never mutates the table or the caller's
    (cached) masters object.

    Rotation itself is REMOVED (2026-08-05, owner decision): the shift an admin
    sets holds every week, so this no longer proves a Friday-elapsed flip (that
    was the whole point of the old `test_prepare_contest_rotates_as_of_effective_
    start`, deleted with this rewrite). The anchor here equals the plan's
    effective start on purpose, so zero Fridays have elapsed either way — the
    assertion doesn't depend on whether `operator_master.rotate_table` still does
    Friday math internally (Task 2) or not."""
    raw = build_sample_bytes()
    so_lines, masters = load_all(io.BytesIO(raw))
    orders = _orders(so_lines)

    friday = date(2026, 1, 9)            # a Friday; also the plan's effective start
    table = {"week_anchor": friday.isoformat(),
             "operators": [
                 {"id": "1", "name": "Alpha", "machines_raw": "CNC 1",
                  "shift": "Second shift", "pinned": False},
                 {"id": "2", "name": "Beta", "machines_raw": "CNC 2",
                  "shift": "First shift", "pinned": True},
             ]}

    setup = svc.prepare_contest(orders, [], masters, _cfg(start=friday),
                                operator_table=table)
    # The table's shifts flow straight through — the contest used the app's
    # operator table, not the workbook's.
    assert _shift_of(setup.masters, "Alpha") == "Second shift"
    assert _shift_of(setup.masters, "Beta") == "First shift"

    # The STORE (here just the passed dict) is never mutated by planning.
    assert table["operators"][0]["shift"] == "Second shift"
    # The cached masters object handed in is never mutated either.
    assert masters.operators == load_all(io.BytesIO(raw))[1].operators


def test_prepare_contest_without_table_uses_masters_operators():
    raw = build_sample_bytes()
    so_lines, masters = load_all(io.BytesIO(raw))
    orders = _orders(so_lines)
    setup = svc.prepare_contest(orders, [], masters, _cfg())
    assert setup.masters is masters                 # no copy when no table


# --------------------------------------------------------------------------- #
# 4. Cloud == local with a payload-carried operator table.
# --------------------------------------------------------------------------- #
def test_cloud_equals_local_with_operator_table():
    raw = build_sample_bytes()
    so_lines, masters = load_all(io.BytesIO(raw))
    orders = _orders(so_lines)
    cfg = _cfg()
    # A table that differs from the workbook (a flipped shift) to prove the
    # overlay actually flows through the payload.
    table = {"week_anchor": date(2026, 1, 2).isoformat(),
             "operators": operator_master.seed_rows_from_masters(masters)}
    table["operators"][0]["shift"] = ("Second shift"
                                      if table["operators"][0]["shift"] == "First shift"
                                      else "First shift")

    payload = svc.build_payload(orders, [], raw, cfg, seed=42,
                                budget_per_candidate=8, operator_table=table)
    payload = json.loads(json.dumps(payload))
    payload["candidates"] = list(optimizer.OVERLAP_CANDIDATES)
    cloud = svc.run_contest(payload, processes=1)

    setup = svc.prepare_contest(orders, [], masters, cfg, operator_table=table)
    n = len(optimizer.sweep_contenders(cfg.overlap_percent,
                                       optimizer.OVERLAP_CANDIDATES))
    local = optimizer.sweep_optimize(setup.target, setup.search_config,
                                     setup.masters, budget_evals=8 * n, seed=42)
    assert cloud["winner_overlap"] == local.overlap_percent
    assert cloud["best"] == local.result.best
    assert cloud["ranks"] == local.result.ranks


def test_payload_carries_operator_table_verbatim():
    raw = build_sample_bytes()
    so_lines, masters = load_all(io.BytesIO(raw))
    orders = _orders(so_lines)
    table = {"week_anchor": "2026-01-02",
             "operators": operator_master.seed_rows_from_masters(masters)}
    payload = svc.build_payload(orders, [], raw, _cfg(), seed=1,
                                operator_table=table)
    payload = json.loads(json.dumps(payload))
    parsed = svc.parse_payload(payload)
    assert len(parsed) == 7
    assert parsed[5] == table


# --------------------------------------------------------------------------- #
# Fingerprints: an operator edit / rotation must move the inputs signature.
# --------------------------------------------------------------------------- #
def test_inputs_signature_reflects_the_operator_table():
    m = _api()
    book_store.save_masters_bytes(build_sample_bytes())
    m._current_masters()                            # seed
    cfg = m._load_plan_config()
    s0 = m._inputs_signature(cfg)

    table = book_store.load_operator_table()
    # Flip an operator's shift → signature must change.
    table["operators"][0]["shift"] = "Second shift"
    book_store.save_operator_table(table)
    assert m._inputs_signature(cfg) != s0

    # Restore → signature returns.
    table = book_store.load_operator_table()
    table["operators"][0]["shift"] = "First shift"
    book_store.save_operator_table(table)
    assert m._inputs_signature(cfg) == s0

    # A pin change also moves it (pins shape future rotations).
    table["operators"][0]["pinned"] = True
    book_store.save_operator_table(table)
    assert m._inputs_signature(cfg) != s0


def test_inputs_signature_ignores_a_pure_anchor_advance():
    """A net-no-op double-Friday catch-up advances week_anchor while every
    shift stays identical. The signature must NOT move — otherwise every
    scheduled tick after such a catch-up would fire a full contest that then
    doesn't apply, forever until the next Apply (reviewer finding). A real
    content change (a flipped shift) must still move it."""
    from datetime import timedelta
    m = _api()
    book_store.save_masters_bytes(build_sample_bytes())
    m._current_masters()                            # seed
    cfg = m._load_plan_config()
    s0 = m._inputs_signature(cfg)

    # Advance ONLY the anchor by 14 days (2 Fridays = content-identical table).
    table = book_store.load_operator_table()
    anchor = date.fromisoformat(table["week_anchor"])
    table["week_anchor"] = (anchor + timedelta(days=14)).isoformat()
    book_store.save_operator_table(table)
    assert m._inputs_signature(cfg) == s0           # anchor alone: EQUAL

    # A genuine shift flip still changes it.
    table = book_store.load_operator_table()
    table["operators"][0]["shift"] = ("Second shift"
                                      if table["operators"][0]["shift"] == "First shift"
                                      else "First shift")
    book_store.save_operator_table(table)
    assert m._inputs_signature(cfg) != s0           # content change: DIFFERS


# --------------------------------------------------------------------------- #
# Scheduled-skip: run when EITHER the book OR the inputs fingerprint differs.
# --------------------------------------------------------------------------- #
def test_scheduled_runs_when_inputs_changed_though_book_same(monkeypatch):
    monkeypatch.setenv("AUTO_OPTIMIZE", "1")
    monkeypatch.setenv("GITHUB_DISPATCH_TOKEN", "manual")
    monkeypatch.setenv("OPTIMIZE_WORKER_SECRET", "s3")
    m = _api()
    book_store.save_masters_bytes(build_sample_bytes())
    book_store.add_orders([Order("SO1", "ITEM_A", "ITEM_A", 10, date(2025, 3, 20))])

    # Applied plan whose book_sig matches now but inputs_sig is stale.
    book_store.save_plan_priority({}, {"saved_at": "t",
                                       "book_sig": m._current_book_sig(),
                                       "inputs_sig": "STALE"})
    starts = []
    monkeypatch.setattr(m, "_start_optimize",
                        lambda *a, **k: starts.append(1))
    assert m._try_start_auto() is True
    assert starts == [1]


def test_scheduled_skips_when_book_and_inputs_both_match(monkeypatch):
    monkeypatch.setenv("AUTO_OPTIMIZE", "1")
    monkeypatch.setenv("GITHUB_DISPATCH_TOKEN", "manual")
    monkeypatch.setenv("OPTIMIZE_WORKER_SECRET", "s3")
    m = _api()
    book_store.save_masters_bytes(build_sample_bytes())
    book_store.add_orders([Order("SO1", "ITEM_A", "ITEM_A", 10, date(2025, 3, 20))])
    cfg = m._load_plan_config()
    book_store.save_plan_priority({}, {"saved_at": "t",
                                       "book_sig": m._current_book_sig(),
                                       "inputs_sig": m._inputs_signature(cfg)})
    starts = []
    monkeypatch.setattr(m, "_start_optimize", lambda *a, **k: starts.append(1))
    assert m._try_start_auto() is False
    assert starts == []


# --------------------------------------------------------------------------- #
# No shift rotation (2026-08-05): the admin's Settings shift holds every week.
# --------------------------------------------------------------------------- #
def test_operator_shift_never_changes_across_a_friday():
    """The bug the director saw: an operator on 1st shift in week 1 showed on 2nd
    shift in week 2. The shift an admin sets must now hold for every week."""
    import datetime
    from engine.new_engine import _plan_config
    from engine.config import Config
    from ppc_engine.domain.resources import Operator as EngOp, Role, Shift
    from ppc_engine.worktime import effective_shift

    start = datetime.date(2026, 8, 5)          # a Wednesday
    cfg = _plan_config(Config(plan_start_date=start))
    op = EngOp(name="Sidhu Singe", role=Role.OPERATOR,
               qualified_machines=frozenset({"CNC1"}), base_shift=Shift.FIRST)

    # Six consecutive weeks, crossing five Fridays.
    shifts = {effective_shift(op, start + datetime.timedelta(days=7 * i), cfg)
              for i in range(6)}
    assert shifts == {Shift.FIRST}, f"shift moved across a Friday: {shifts}"
