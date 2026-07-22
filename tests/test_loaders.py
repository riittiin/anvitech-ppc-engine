"""Loader tests — expected counts and the known non-blocking data quirks."""
import openpyxl

from engine.loaders import (
    load_all, normalize_resource_id, parse_date, parse_resource_candidates, _num,
    _validate, _load_so_lines,
)
from engine.models import Masters, Routing, Process
from datetime import date


def test_num_rejects_non_finite_values():
    # A cell typed 'nan'/'inf' (or a NaN/inf float) would make occupancy non-finite
    # and hang the scheduler — the loader must drop it to None.
    assert _num("nan") is None
    assert _num("inf") is None
    assert _num("infinity") is None
    assert _num(float("nan")) is None
    assert _num("5.5") == 5.5
    assert _num(10) == 10.0


def test_counts(loaded):
    so_lines, masters = loaded
    assert len(so_lines) == 3                       # 3 schedulable SO lines
    assert len(masters.routings) == 2               # 2 item codes (SAMP-A/B)
    # Real (non-provisional) machines in the sample master: CNC1, CNC2, VMC1,
    # BS1, MI1, MW1 = 6.
    real = [m for m in masters.machines.values() if not m.provisional]
    assert len(real) == 6


def test_pending_master_data_is_nonblocking(loaded):
    _, masters = loaded
    pending = [r for r in masters.report if r["kind"] == "PENDING_MASTER_DATA"]
    refs = {r["ref"] for r in pending}
    # CNC9 is referenced by a routing but not in the master -> provisional.
    assert "CNC9" in refs
    assert masters.machines["CNC9"].provisional is True


def test_no_routing_count_is_zero(loaded):
    # Every SO item code in the sample has a routing -> no NO_ROUTING.
    _, masters = loaded
    assert [r for r in masters.report if r["kind"] == "NO_ROUTING"] == []


def test_calendar(loaded):
    _, masters = loaded
    assert masters.calendar.weekly_off_weekday == 3          # Thursday
    assert date(2025, 1, 1) in masters.calendar.holidays     # New Year holiday
    assert not masters.calendar.is_working_day(date(2025, 3, 13))  # a Thursday


def test_resource_normalization_matches_spaced_and_compact():
    # The master's 'CNC 4' and a routing's 'CNC4' must collapse onto one key.
    assert normalize_resource_id("CNC 4") == normalize_resource_id("CNC4") == "CNC4"
    assert normalize_resource_id("VMC 1") == "VMC1"


def test_parse_date_handles_string_and_datetime():
    from datetime import datetime
    assert parse_date("28/03/2025") == date(2025, 3, 28)
    assert parse_date(datetime(2025, 3, 28)) == date(2025, 3, 28)


def test_parse_date_is_strictly_day_first():
    # Day-first slash/dash + ISO are accepted; an ambiguous date is read DAY-first.
    assert parse_date("03/04/2025") == date(2025, 4, 3)     # 3 April, not 4 March
    assert parse_date("03-04-2025") == date(2025, 4, 3)
    assert parse_date("2025-04-03") == date(2025, 4, 3)
    # A month-first-only US string (day > 12 in the 2nd slot) is now REJECTED,
    # not silently coerced — the parser is strictly day-first.
    assert parse_date("12/25/2025") is None


# --------------------------------------------------------------------------- #
# Alternative ("preferred") machines: a cell like "CNC3/CNC6" = either machine
# --------------------------------------------------------------------------- #
def test_parse_resource_candidates_splits_alternatives():
    assert parse_resource_candidates("CNC3/CNC6") == ["CNC3", "CNC6"]
    assert parse_resource_candidates("CNC 3 / CNC 6") == ["CNC3", "CNC6"]   # spaces
    assert parse_resource_candidates("CNC6/CNC3") == ["CNC6", "CNC3"]       # order kept
    assert parse_resource_candidates("CNC3 or CNC6") == ["CNC3", "CNC6"]    # 'or'


def test_parse_resource_candidates_single_and_empty():
    assert parse_resource_candidates("VMC2") == ["VMC2"]
    assert parse_resource_candidates("") == []
    assert parse_resource_candidates(None) == []
    assert parse_resource_candidates("   ") == []
    assert parse_resource_candidates("CNC6/CNC6") == ["CNC6"]               # dedupe


def test_alternative_cells_register_each_machine_not_a_merged_id(loaded):
    _, masters = loaded
    # The sample's "CNC1/CNC2" alternative cell must register each machine
    # separately, never a merged id.
    for bogus in ("CNC1CNC2", "CNC2CNC1"):
        assert bogus not in masters.machines
    assert "CNC1" in masters.machines and not masters.machines["CNC1"].provisional
    assert "CNC2" in masters.machines and not masters.machines["CNC2"].provisional


# --------------------------------------------------------------------------- #
# SO list: header-driven, position-independent (the owner's ERP export shifted
# every SO column and NOT uniformly — a fixed-index reader silently reads the
# wrong columns and drops every row as BAD_DELIVERY_DATE).
# --------------------------------------------------------------------------- #
def _build_so_workbook(header_by_col: dict, data_rows: list):
    """Build a tiny workbook with just the SO sheet. ``header_by_col`` maps
    0-based column index -> header text (row 1). ``data_rows`` is a list of
    {0-based col: value} dicts, one per data row (from row 2)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Order (SO) list"
    for col, text in header_by_col.items():
        ws.cell(row=1, column=col + 1, value=text)
    for r, row in enumerate(data_rows, start=2):
        for col, value in row.items():
            ws.cell(row=r, column=col + 1, value=value)
    return wb


# Mirrors the real Test5.xlsx layout: SONo=14, Customer Name=17,
# Sales Item Code=28, Sales Item Name=29, SO Qty=30, SO Delivery Date=32,
# Pend SO Qty=35 (0-based) — shifted right from the old fixed indices, and
# NOT uniformly shifted.
SHIFTED_HEADERS = {
    14: "SONo", 17: "Customer Name", 28: "Sales Item Code",
    29: "Sales Item Name", 30: "SO Qty", 32: "SO Delivery Date",
    35: "Pend SO Qty",
}


def test_so_lines_survive_a_column_shift():
    # THIS is the regression: the old fixed-index reader assumed
    # so_no=5, item_code=19, qty=21, delivery=23 — all wrong here — and would
    # read the delivery date from a column holding text, dropping every row.
    wb = _build_so_workbook(SHIFTED_HEADERS, [
        {14: "SO-1", 17: "Cust A", 28: "ITEM-1", 29: "Widget", 30: 5,
         32: "10/08/2026", 35: 5},
        {14: "SO-2", 17: "Cust B", 28: "ITEM-2", 29: "Gadget", 30: 8,
         32: "12/08/2026", 35: 3},
    ])
    masters = Masters()
    so_lines = _load_so_lines(wb, masters)
    assert len(so_lines) == 2
    assert so_lines[0].so_no == "SO-1"
    assert so_lines[0].item_code == "ITEM-1"
    assert so_lines[0].qty == 5.0
    assert so_lines[0].delivery_date == date(2026, 8, 10)
    assert so_lines[0].customer == "Cust A"
    assert so_lines[0].pending_qty == 5.0
    assert so_lines[1].so_no == "SO-2"
    assert so_lines[1].item_code == "ITEM-2"
    assert so_lines[1].qty == 8.0
    assert so_lines[1].delivery_date == date(2026, 8, 12)


def test_so_qty_not_shadowed_by_pend_so_qty_to_its_left():
    # Token collision: "SO Qty"->"soqty" is a SUBSTRING of "Pend SO Qty"->
    # "pendsoqty". _locate_table binds the leftmost substring match, so if
    # Pend SO Qty sits LEFT of SO Qty the qty would silently read the pending
    # value (999) instead of the ordered quantity (5) — a silently-wrong order
    # qty, worse than a hard failure. exact_priority makes "soqty" win by exact
    # match regardless of column order.
    headers = {
        14: "SONo", 17: "Customer Name", 28: "Sales Item Code",
        29: "Sales Item Name", 30: "Pend SO Qty", 31: "SO Qty",
        32: "SO Delivery Date",
    }
    wb = _build_so_workbook(headers, [
        {14: "SO-1", 17: "Cust A", 28: "ITEM-1", 29: "Widget",
         30: 999, 31: 5, 32: "10/08/2026"},
    ])
    masters = Masters()
    so_lines = _load_so_lines(wb, masters)
    assert len(so_lines) == 1
    assert so_lines[0].qty == 5.0            # SO Qty, NOT Pend SO Qty (999)
    assert so_lines[0].pending_qty == 999.0


def test_so_lines_tolerate_no_remarks_column():
    # The real Test5.xlsx has NO "Remarks" column at all — it must be optional.
    wb = _build_so_workbook(SHIFTED_HEADERS, [
        {14: "SO-1", 17: "Cust A", 28: "ITEM-1", 29: "Widget", 30: 5,
         32: "10/08/2026", 35: 5},
    ])
    masters = Masters()
    so_lines = _load_so_lines(wb, masters)
    assert len(so_lines) == 1
    assert so_lines[0].remarks == ""


def test_so_lines_missing_required_column_reports_and_returns_empty():
    # Omit "SO Delivery Date" entirely -> can't find the header row by name.
    headers = {k: v for k, v in SHIFTED_HEADERS.items() if v != "SO Delivery Date"}
    wb = _build_so_workbook(headers, [
        {14: "SO-1", 17: "Cust A", 28: "ITEM-1", 29: "Widget", 30: 5, 35: 5},
    ])
    masters = Masters()
    so_lines = _load_so_lines(wb, masters)
    assert so_lines == []
    kinds = [r["kind"] for r in masters.report]
    assert "MISSING_SO_COLUMNS" in kinds


def test_bad_qty_is_reported_not_silently_zeroed():
    # A present-but-unparseable SO Qty (e.g. "5 Nos") must be reported
    # (BAD_QTY), not silently coerced to 0.0 with no warning.
    wb = _build_so_workbook(SHIFTED_HEADERS, [
        {14: "SO-1", 17: "Cust A", 28: "ITEM-1", 29: "Widget", 30: "5 Nos",
         32: "10/08/2026", 35: 5},
    ])
    masters = Masters()
    so_lines = _load_so_lines(wb, masters)
    assert len(so_lines) == 1
    assert so_lines[0].qty == 0.0
    bad_qty = [r for r in masters.report if r["kind"] == "BAD_QTY"]
    assert len(bad_qty) == 1
    assert bad_qty[0]["ref"] == "SO-1"
    assert "5 Nos" in bad_qty[0]["message"]


def test_blank_qty_stays_silent_zero_no_report():
    # A truly empty qty cell keeps the existing (silent) behaviour — only a
    # present-but-unparseable value is reported.
    wb = _build_so_workbook(SHIFTED_HEADERS, [
        {14: "SO-1", 17: "Cust A", 28: "ITEM-1", 29: "Widget",
         32: "10/08/2026", 35: 5},   # no col 30 (SO Qty) at all
    ])
    masters = Masters()
    so_lines = _load_so_lines(wb, masters)
    assert len(so_lines) == 1
    assert so_lines[0].qty == 0.0
    assert [r for r in masters.report if r["kind"] == "BAD_QTY"] == []


def test_os_is_not_registered_as_a_machine():
    # An outsourced step (Allotted = OS) must NOT create a phantom 'OS' machine
    # or a PENDING_MASTER_DATA report — OS is a sentinel, not a resource.
    proc = Process(seq=1, name="CNC OS", cycle_time=7200, total_time=None,
                   suggested_machine=None, allotted_machine="OS")
    masters = Masters(routings={"X": Routing(item_code="X", description="", customer="",
                                             rm_type="", moq=None, processes=[proc])})
    _validate(masters, [])
    assert normalize_resource_id("OS") not in masters.machines
    assert not any(r["ref"] == "OS" for r in masters.report)
