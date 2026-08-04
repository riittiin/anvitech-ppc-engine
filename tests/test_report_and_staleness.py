"""Regressions for the 2026-07-15 live findings (bug 2 + bug 3):

* The "N orders without routing" banner showed items that are NOT in the order
  book (a stale loader report from the stored workbook's own SO sheet — the live
  site showed 5 ghosts while all 54 real orders planned fine). The plan report's
  NO_ROUTING rows must be derived from the CURRENT book.

* An applied optimization silently kept replaying after the masters or plan
  settings changed (owner re-uploaded an edited workbook after Apply), so "the
  same Deep-400 run" appeared to give different numbers on different days. The
  applied result now carries a fingerprint of the inputs it was computed on, and
  /run's optimize_meta flags `inputs_changed` so the UI can say why.
"""
from datetime import date

import pytest

pytest.importorskip("fastapi")
from fastapi.testclient import TestClient  # noqa: E402

from engine import book_store
from engine.models import Order
from tests.sample_workbook import build_sample_bytes, build_workbook, ITEM_A, ITEM_B


def _api():
    import importlib
    import api.main as m
    importlib.reload(m)
    return m


def _seed_book(n_orders=2):
    book_store.save_masters_bytes(build_sample_bytes())
    items = [ITEM_A, ITEM_B]
    book_store.add_orders([
        Order(f"SO{i+1}", items[i % 2], items[i % 2], 10 + 5 * i, date(2025, 3, 20 + i))
        for i in range(n_orders)])


def _admin_client(m):
    c = TestClient(m.app)
    c.post("/login", data={"username": "anvitech", "password": "1930rail"})
    return c


# --------------------------------------------------------------------------- #
# Bug 3 — NO_ROUTING report reflects the CURRENT book, not the stored workbook
# --------------------------------------------------------------------------- #
def test_plan_report_no_routing_is_derived_from_the_book():
    m = _api()
    _seed_book()
    # An order in the book whose item has no routing in the masters -> reported.
    book_store.add_orders([Order("SO9", "GHOSTLESS-ITEM", "X", 5, date(2025, 3, 25))])
    payload = m._plan(m._load_plan_config())
    rows = payload["report"]["rows"]
    no_routing = [r for r in rows if r[0] == "NO_ROUTING"]
    assert ["NO_ROUTING", "GHOSTLESS-ITEM"] == [no_routing[0][0], no_routing[0][1]]
    # ... and ONLY book items appear (nothing from the workbook's own SO sheet).
    book_items = {ITEM_A, ITEM_B, "GHOSTLESS-ITEM"}
    assert all(r[1] in book_items for r in no_routing)


def test_plan_report_shows_no_ghost_no_routing_rows():
    """All book items have routings -> zero NO_ROUTING rows, whatever the stored
    workbook's SO sheet contains (the live '5 orders without routing' ghosts)."""
    m = _api()
    _seed_book()
    payload = m._plan(m._load_plan_config())
    assert [r for r in payload["report"]["rows"] if r[0] == "NO_ROUTING"] == []


# --------------------------------------------------------------------------- #
# 2026-07-18 — the upload response must show the FILE's own dropped (no-routing)
# item codes; /run|/gantt|/report stay book-scoped (the ghost fix above).
# --------------------------------------------------------------------------- #
def test_upload_report_contains_no_routing_for_the_files_dropped_item():
    """A SO line whose item has no routing is dropped before it ever reaches the
    order book, so the book-scoped /run report can never show it. The upload
    endpoint must surface it directly from the loader's own report so the admin
    knows which item code to add to the Item's process Master."""
    import io

    m = _api()
    from api import auth

    client = TestClient(m.app)
    accts = auth._accounts()
    admin = next(u for u, a in accts.items() if a["role"] == auth.ADMIN)
    pwd = accts[admin]["password"]
    assert client.post("/login", data={"username": admin, "password": pwd}).status_code == 200

    wb = build_workbook()
    ws = wb["Sales Order (SO) list"]
    row = ws.max_row + 1
    # Same column layout sample_workbook.build_workbook uses for SO rows
    # (0-based col -> 1-based openpyxl col): SONo=6, Customer=9, Item code=20,
    # Item name=21, Qty=22, Delivery=24, Remarks=25, Pend Qty=28.
    ws.cell(row=row, column=6, value="SO-404")
    ws.cell(row=row, column=9, value="ALFA LAVAL")
    ws.cell(row=row, column=20, value="NOROUTE-ITEM")
    ws.cell(row=row, column=21, value="No Routing Item")
    ws.cell(row=row, column=22, value=3)
    ws.cell(row=row, column=24, value=date(2025, 3, 30))
    ws.cell(row=row, column=25, value="")
    ws.cell(row=row, column=28, value=3)
    buf = io.BytesIO()
    wb.save(buf)

    xlsx_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp = client.post("/upload", files={"file": ("t.xlsx", buf.getvalue(), xlsx_mime)})
    assert resp.status_code == 200
    report = resp.json()["report"]
    no_routing = [r for r in report["rows"] if r[0] == "NO_ROUTING"]
    assert any(r[1] == "NOROUTE-ITEM" for r in no_routing), no_routing


def test_report_after_upload_unit_keeps_masters_report_no_routing():
    """`_report_after_upload` returns masters.report AS-IS (no book-scoping)."""
    m = _api()
    from engine.models import Masters

    masters = Masters()
    masters.add_report("NO_ROUTING", "SOME-ITEM",
                        "SO item 'SOME-ITEM' has no routing in Item's process "
                        "Master; order skipped (cannot schedule without a recipe)")
    table = m._report_after_upload(masters)
    no_routing = [r for r in table["rows"] if r[0] == "NO_ROUTING"]
    assert ["NO_ROUTING", "SOME-ITEM"] == [no_routing[0][0], no_routing[0][1]]


# --------------------------------------------------------------------------- #
# Bug 2 — applied optimization fingerprints its inputs; /run flags changes
# --------------------------------------------------------------------------- #
def test_optimize_meta_flags_inputs_changed_when_settings_change():
    m = _api()
    _seed_book()
    st = m._start_optimize(budget_evals=10, label="quick", background=False)
    assert st["state"] == "done"
    m._optimize_apply()
    saved = book_store.load_plan_priority()
    assert saved["meta"].get("inputs_sig"), "applied optimization must carry a fingerprint"

    cfg = m._load_plan_config()
    meta_same = m._plan(cfg)["optimize_meta"]
    assert meta_same["active"] is True
    assert meta_same["inputs_changed"] is False        # same inputs -> no warning

    from dataclasses import replace
    changed = replace(cfg, overlap_percent=60)          # the owner edits Settings
    meta_changed = m._plan(changed)["optimize_meta"]
    assert meta_changed["inputs_changed"] is True


def test_finalize_recomputes_best_from_local_replay():
    """The reported/stored 'best' must be the winner's ranks replayed LOCALLY, not the
    number the contest (cloud worker) reported — so the panel and the applied plan
    always agree (closes the 52.5-promised / 55.6-applied gap)."""
    import time
    m = _api(); _seed_book()
    m._start_optimize(budget_evals=10, label="quick", background=False)
    res = m._OPTIMIZE["result"]
    ranks, ov = res["ranks"], res["best_overlap"]
    baseline = m._OPTIMIZE["baseline"]
    honest = m._metrics_for_ranks(ranks, ov)          # what the plan really achieves
    assert honest and honest["makespan_days"] is not None

    # Finalize a NEW job with a BOGUS best (as if the contest over-reported makespan).
    m._OPTIMIZE.update(job_id="jobZ", started_mono=time.monotonic(),
                       searched_book_sig="x", searched_inputs_sig="y", auto=False)
    bogus = dict(honest); bogus["makespan_days"] = 1.0
    m._finalize_optimize("jobZ", m._load_plan_config(), baseline, "quick",
                         winner_overlap=ov, ranks=ranks, best=bogus,
                         evals=10, table=[], cancelled=False)
    stored = m._OPTIMIZE["result"]["best"]
    assert stored["makespan_days"] == honest["makespan_days"], "best was not recomputed locally"
    assert stored["makespan_days"] != 1.0


def test_optimize_meta_surfaces_achieved_vs_target_no_divergence():
    """A freshly applied optimization: the live plan achieves what it targeted (the
    ranks replay to the same plan), so optimize_meta carries achieved + target and
    does NOT flag divergence."""
    m = _api(); _seed_book()
    assert m._start_optimize(budget_evals=10, label="quick", background=False)["state"] == "done"
    m._optimize_apply()
    meta = m._plan(m._load_plan_config())["optimize_meta"]
    assert meta["target"]["makespan_days"] is not None
    assert meta["achieved"]["makespan_days"] is not None
    assert meta["diverged"] is False


def test_optimize_meta_flags_divergence_when_plan_misses_target():
    """When the live plan is meaningfully worse than the applied optimization's target
    — a gap the inputs signature can't see (a code deploy mid-contest, a cloud/local
    parity gap) — optimize_meta flags `diverged` so the UI can warn to re-optimize.
    This is the 2026-07-25 'optimizer said 52, applied plan is 56' wiring gap."""
    m = _api(); _seed_book()
    m._start_optimize(budget_evals=10, label="quick", background=False)
    m._optimize_apply()
    # Rewrite the applied target to an impossibly-good makespan (as if the number the
    # panel showed came from a state the live plan no longer reproduces).
    saved = book_store.load_plan_priority()
    meta = dict(saved["meta"]); meta["best"] = {"makespan_days": 0.0, "total_late_days": 0}
    book_store.save_plan_priority(saved["ranks"], meta)
    m._PLAN_CACHE.update(key=None, result=None)
    om = m._plan(m._load_plan_config())["optimize_meta"]
    assert om["target"]["total_late_days"] == 0
    assert om["achieved"]["total_late_days"] >= 5      # the live plan is far worse
    assert om["diverged"] is True


def test_optimize_meta_flags_inputs_changed_when_masters_change():
    m = _api()
    _seed_book()
    m._start_optimize(budget_evals=10, label="quick", background=False)
    m._optimize_apply()
    cfg = m._load_plan_config()
    assert m._plan(cfg)["optimize_meta"]["inputs_changed"] is False

    # The owner re-uploads an edited workbook (masters latest-wins).
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(build_sample_bytes()))
    ws = wb["Machine master"]
    ws.cell(row=2, column=1, value=ws.cell(row=2, column=1).value)  # touch nothing...
    ws.cell(row=ws.max_row + 1, column=1, value="CNC99")            # ...then add a machine
    buf = io.BytesIO(); wb.save(buf)
    book_store.save_masters_bytes(buf.getvalue())
    m._MASTERS_CACHE.update(key=None, masters=None)                 # what /upload does

    assert m._plan(cfg)["optimize_meta"]["inputs_changed"] is True


def test_schedule_neutral_settings_do_not_flag_staleness():
    """Balance-operator-workload never changes timing, and expedite is forced off
    under an applied optimization — toggling either must NOT cry 'stale'."""
    m = _api()
    _seed_book()
    m._start_optimize(budget_evals=10, label="quick", background=False)
    m._optimize_apply()
    from dataclasses import replace
    cfg = m._load_plan_config()
    for tweak in (replace(cfg, balance_operator_load=True),
                  replace(cfg, expedite_window_min=45)):
        assert m._plan(tweak)["optimize_meta"]["inputs_changed"] is False


# --------------------------------------------------------------------------- #
# 2026-08-04 — the applied optimization records the delivery dates it was
# computed against, so a later re-import that moves a date flags staleness.
# --------------------------------------------------------------------------- #
def test_optimize_meta_flags_a_changed_delivery_date(monkeypatch):
    """After an optimization is applied, changing an order's delivery date must
    mark the applied plan stale so the admin knows to run Start deep search."""
    m = _api()
    admin = _admin_client(m)
    _seed_book()

    # Save an applied optimization whose recorded dates match the book.
    dates = m._delivery_dates()
    assert dates, "the seeded book should have at least one order"
    key = sorted(dates)[0]
    book_store.save_plan_priority({key: 0}, {"saved_at": "2026-08-04T10:00:00",
                                              "dates": dict(dates)})

    meta = admin.post("/run", json={}).json()["optimize_meta"]
    assert meta["dates_changed"] is False
    assert meta["dates_changed_count"] == 0

    # Move that order's delivery date.
    so_no, item_code = key.split(m.KEY_SEP)
    active = book_store.load_active_orders()
    order = active[(so_no, item_code)]
    import dataclasses
    import datetime
    book_store.add_orders([dataclasses.replace(
        order, delivery_date=order.delivery_date + datetime.timedelta(days=30))])

    meta = admin.post("/run", json={}).json()["optimize_meta"]
    assert meta["dates_changed"] is True
    assert meta["dates_changed_count"] == 1


def test_optimize_meta_ignores_orders_the_optimization_never_saw(monkeypatch):
    """Only orders present in BOTH the applied snapshot and the current book are
    compared. A newly uploaded or newly completed order is normal traffic, not a
    reason to tell the admin their optimization is stale."""
    m = _api()
    admin = _admin_client(m)
    _seed_book()

    dates = m._delivery_dates()
    key = sorted(dates)[0]
    stale = dict(dates)
    stale["SO_GONE" + m.KEY_SEP + "ITEM_GONE"] = "2020-01-01"   # not in the book
    book_store.save_plan_priority({key: 0}, {"saved_at": "2026-08-04T10:00:00",
                                              "dates": stale})

    meta = admin.post("/run", json={}).json()["optimize_meta"]
    assert meta["dates_changed"] is False
    assert meta["dates_changed_count"] == 0
