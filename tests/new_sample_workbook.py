"""A synthetic workbook that is fully compatible with the NEW engine.

Unlike ``sample_workbook`` (built for the classic engine and deliberately under-staffed),
this one satisfies the new engine's stricter requirements so the new-engine tests can
actually run it:
  * an explicit **Role** column on the operator sheet (operator / helper / inspector),
  * **every machine used by a routing is staffed** by a qualified operator of the matching
    role (operator↔machining, helper↔manual, inspector↔inspection),
  * routings exercise all step kinds: machining, manual, inspection, outsourced (OS),
    and a terminal DISPATCH milestone,
  * an item with **two machining ops** (for the per-process feedback test).

Same sheet layout/format the real workbook uses, so the new loaders parse it natively.
"""

import datetime
import io

import openpyxl

MACHINES = [
    ("CNC lathe", "CNC1", 250, 19.5),
    ("CNC lathe", "CNC2", 250, 19.5),
    ("Vertical Machining center", "VMC1", 500, 19.5),
    ("Manual Inspection", "MI1", 150, 9.5),
    ("Manual Washing", "MW1", 80, 9.5),
]

# name, role, preferred machines, shift
OPERATORS = [
    ("Alpha", "operator", "CNC1, CNC2, VMC1", "First shift"),
    ("Bravo", "operator", "CNC1, CNC2, VMC1", "Second shift"),
    ("Charlie", "helper", "MW1", "First shift"),
    ("Delta", "inspector", "MI1", "First shift"),
]

ITEM_A, ITEM_B = "NEW-A-01", "NEW-B-01"
SO1, SO2 = "NSO-001", "NSO-002"


def _set(ws, row, mapping):
    for ci, val in mapping.items():
        ws.cell(row=row, column=ci + 1, value=val)


def build_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Machine master")
    ws.append(["Machine Type", "Machine No", "Hr Rate (Rs)", "Available Hrs/Day"])
    for r in MACHINES:
        ws.append(list(r))

    # Operator & shift Master — WITH a Role column (col B), shift-timings mini-table to the side.
    ws = wb.create_sheet("Operator & shift Master")
    ws.append(["Operator Name", "Role", "Preferred Machines", "Shift", None,
               "Shift", "Start", "End", "Notes"])
    shifts = [("First shift", "08:00", "19:00", ""),
              ("Second shift", "19:00", "05:00", "next day")]
    for i in range(max(len(OPERATORS), len(shifts))):
        row = [None] * 9
        if i < len(OPERATORS):
            row[0], row[1], row[2], row[3] = OPERATORS[i]
        if i < len(shifts):
            row[5], row[6], row[7], row[8] = shifts[i]
        ws.append(row)

    ws = wb.create_sheet("Weekly off & holiday master")
    ws.append(["Category", "Name", "Day / Date"])
    ws.append(["Weekly Off", "", "Every Thursday"])

    # Item's process Master (header row 2; data row 3+). Five process blocks of 5 cols.
    ws = wb.create_sheet("Item's process Master")
    _set(ws, 1, {5: "Order Type", 6: "Raw material data", 11: "Process sequence"})
    hdr = {0: "Master Sr NO", 1: "Customer", 2: "Item Description", 3: "Item code",
           6: "RM type", 10: "MOQ/Batch qty"}
    for p in range(5):  # Process 1..5 at columns 12,17,22,27,32
        base = 12 + p * 5
        hdr[base] = f"Process {p+1}"
        hdr[base + 1] = f"Process {p+1} cycle time"
        hdr[base + 2] = f"Process {p+1} Total time"
        hdr[base + 3] = f"Process {p+1} Suggested M/c"
        hdr[base + 4] = f"Process {p+1} Allotted M/c"
    _set(ws, 2, hdr)
    # Item A: CNC -> VMC -> INSPECTION -> DISPATCH  (two machining ops)
    _set(ws, 3, {0: 1, 1: "ACME", 2: "NEW RING A", 3: ITEM_A, 6: "Dia 20 SS", 10: 50,
                 12: "CNC FIRST SIDE", 13: 3, 14: 3, 15: "CNC1/CNC2",
                 17: "VMC FIRST SIDE", 18: 4, 19: 4, 20: "VMC1",
                 22: "INSPECTION", 23: 2, 24: 2, 25: "MI1",
                 27: "DISPATCH", 28: 0, 29: 0})
    # Item B: BANDSAW OS -> CNC -> WASHING -> DISPATCH  (OS + machining + manual)
    _set(ws, 4, {0: 2, 1: "ACME", 2: "NEW PIN B", 3: ITEM_B, 6: "Dia 14 SS", 10: 100,
                 12: "BANDSAW OS", 13: 240, 14: 240,
                 17: "CNC SECOND SIDE", 18: 4, 19: 4, 20: "CNC1/CNC2",
                 22: "WASHING", 23: 2, 24: 2, 25: "MW1",
                 27: "DISPATCH", 28: 0, 29: 0})

    ws = wb.create_sheet("Sales Order (SO) list")
    _set(ws, 1, {5: "SONo", 8: "Customer Name", 19: "Sales Item Code",
                 20: "Sales Item Name", 21: "SO Qty", 23: "SO Delivery Date",
                 24: "Remarks", 27: "Pend SO Qty"})
    _set(ws, 2, {5: SO1, 8: "ACME", 19: ITEM_A, 20: "NEW RING A",
                 21: 50, 23: datetime.date(2025, 3, 20), 24: "", 27: 50})
    _set(ws, 3, {5: SO2, 8: "ACME", 19: ITEM_B, 20: "NEW PIN B",
                 21: 100, 23: datetime.date(2025, 3, 25), 24: "", 27: 100})
    return wb


def build_new_sample_bytes() -> bytes:
    buf = io.BytesIO()
    build_workbook().save(buf)
    return buf.getvalue()
