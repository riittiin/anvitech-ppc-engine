"""Tolerant Excel access — open a workbook, find sheets, and read tables by header.

Wraps openpyxl so the rest of the loaders never touch cell coordinates directly.
Two tolerances baked in (LESSONS.md):
  - sheet names are matched ignoring case and surrounding whitespace (several real
    sheet names have trailing spaces), and
  - columns are found by header NAME, never by fixed index.
"""

from __future__ import annotations

from dataclasses import dataclass

import openpyxl

from ppc_engine.loaders.normalize import _norm_header, match_header


def open_workbook(path):
    """Open a workbook read-only (values only — no formulas, no writing)."""
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def find_sheet(wb, *wanted_names):
    """Return the worksheet whose name matches any of ``wanted_names``.

    Matching ignores case and surrounding whitespace, so ``"PPC logics "`` (trailing
    space) or a differently-cased name still resolves. Raises KeyError if none match.
    """
    norm_wanted = [_norm_header(n) for n in wanted_names]
    for ws in wb.worksheets:
        if _norm_header(ws.title) in norm_wanted:
            return ws
    raise KeyError(f"no sheet matching {wanted_names!r}; have {wb.sheetnames!r}")


def rows_of(ws) -> list[tuple]:
    """All rows of a worksheet as a list of value tuples."""
    return list(ws.iter_rows(values_only=True))


def locate_header_row(rows: list[tuple], *key_names: str, search_depth: int = 6) -> int:
    """Find the header row: the first of the top ``search_depth`` rows that contains
    one of ``key_names`` as a column header.

    This tolerates junk/merged cells above the real header (e.g. the Item's process
    Master has a partial row 1 above the true header on row 2).
    """
    for i in range(min(search_depth, len(rows))):
        if match_header(rows[i], *key_names) is not None:
            return i
    raise KeyError(f"could not find a header row containing any of {key_names!r}")


@dataclass
class Table:
    """A sheet viewed as a header + data rows, with column lookup by name."""

    header: tuple
    data_rows: list[tuple]

    @classmethod
    def from_rows(cls, rows: list[tuple], header_index: int) -> "Table":
        return cls(header=rows[header_index], data_rows=rows[header_index + 1 :])

    def col(self, *names: str) -> int | None:
        """Column index for the first matching header name (exact-first)."""
        return match_header(self.header, *names)

    @staticmethod
    def get(row: tuple, index: int | None):
        """Value at ``index`` in ``row``, or None if the index is missing/out of range."""
        if index is None or index >= len(row):
            return None
        return row[index]
